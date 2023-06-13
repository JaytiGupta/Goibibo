import os
import openpyxl


class ExcelFunctions:

    def __init__(self, path):
        self.path = path
        # to load the workbook with its path
        self.bk = openpyxl.load_workbook(self.path)
        # to identify active worksheet
        self.s = self.bk.active

    def write(self, row, column, data):
        # to identify the cell
        c = self.s.cell(row = row, column = column)
        # to write a value in that cell
        c.value = data
        # to save the workbook in location
        self.bk.save(self.path)

    def read(self, row, column):
        c = self.s.cell(row=row, column=column)
        return c.value


if __name__ == "__main__":
    # a = ExcelFunctions(os.environ['USERPROFILE']+'\\Documents\\Automation\\Excel_Python\\Python.xlsx')
    # a.write(3, 4, 'abc')
    # a.write(3, 8, 25)
    # print(a.read(3, 8))
    # print(a.read(3, 4))
    # print(os.environ['USERPROFILE']
    print(os.path.dirname(os.path.abspath(__file__)))

